"""
.
"""
import unittest
import extract_info as hf
import extraction_file as ef
from openpyxl import load_workbook
import os

class TestExtractionFile(unittest.TestCase):
    """
    .
    """
    def test_get_last_total(self):
        """
        .
        """
        file_name = 'statements/FirstD_12_2023.csv'
        bank, month, year = hf.extract_name(file_name.split('.')[0])

        statement = ef.FinanacialManager(bank, year, month)

    def test_get_last_cell(self):
        """
        .
        """
        workbook = load_workbook(filename='finance/2023.xlsx')
        print(workbook.sheetnames)
        sheet = workbook['11']

        for row in sheet.iter_rows(min_row=sheet.max_row-1, max_row=sheet.max_row-1, min_col= 5, max_col= 5, values_only=False):
            for cell in row:
                workbook.close()
                # Returns the sheet the cell is from, the cell value,
                # the value in that cell,the whole cell tuple, veriable type
                print(cell.parent.title, cell.coordinate, cell.value, cell, type(cell)) # String

    def test_new_anything(self):
        """
        Test when you add the very first year in the file finance
        """
        for file_name in ['statements/Barclay_11_2023.csv', 'statements/FirstDirect_11_2023.csv',
                          'statements/Barclay_12_2023.csv', 'statements/FirstDirect_12_2023.csv']:
            print(file_name)
            bank, month, year = hf.extract_name(file_name.split('.')[0])

            statement = ef.FinanacialManager(bank, year, month)
            data = statement.clean(file_name)
            statement.tally_account(data)

    def test_closet_file(self):
        """
        .
        """
        for file_name in ['statements/Barclay_11_2023.csv', 'statements/FirstDirect_11_2023.csv',
                          'statements/Barclay_12_2023.csv', 'statements/FirstDirect_12_2023.csv']:
            # file_name = 'statements/TestA_11_2023.csv'
            bank, month, year = hf.extract_name(file_name.split('.')[0])

            statement = ef.FinanacialManager(bank, year, month)
            data = statement.clean(file_name)
            statement.tally_account(data)

        # Test the update
        for file_name in ['statements/Barclay_01_2024.csv', 'statements/FirstDirect_01_2024.csv',
                          'statements/Barclay_02_2024.csv', 'statements/FirstDirect_02_2024.csv']:
            bank, month, year = hf.extract_name(file_name.split('.')[0])

            statement = ef.FinanacialManager(bank, year, month)
            data = statement.clean(file_name)
            statement.tally_account(data)
            statement.update_sheets()

    def test_order_statements(self):
        """
        .
        """
        statement = [f'statements/{file}' for file in os.listdir('statements/')]
        sorted_files = sorted(statement,
                              key=lambda x: (int(x.split('_')[-1].split('.')[0]), int(x.split('_')[-2]))
                              )
        print(sorted_files)


if __name__ == '__main__':
    unittest.main()

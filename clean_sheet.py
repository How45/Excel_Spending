"""
openpyxl : workbook, PattenFill, Font, Alignment
"""
import json
from icecream import ic
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from extract_info import memo_extraction


def clean_cells(output_file:str, month:str, given_idx: int) -> None:
    """
    Cleans each row of workbook sheet.

    The function will iterate though each row of the workbook sheet;
    - Change width of column
    - Add colour to category
    - Convert all currancy to £, at 2 decimal place.

    Parameters:
    ----------
    output_file : str
    File location. This should be finance/year.xlsx
    month : str
    The month its searching in the workbook. i.e 1 (Jan) or 2 (Feb)
    data : pd.DataFrame
    This is the pandas dataframe of all the colour, so it can fill into the sheet
    given_idx : int
    Where to start in the spreadsheet, either at the end or could be at the start
    """
    workbook = load_workbook(filename=output_file)
    sheet = workbook[month]

    sheet.column_dimensions['A'].width = 10.50 # 2.10
    sheet.column_dimensions['C'].width = 13.05 # 2.61
    sheet.column_dimensions['D'].width = 18.10 # 3.62
    sheet.column_dimensions['E'].width = 12.10 # 2.42

    # If starting index 0 (+2) if its a continuation of a line +1
    if given_idx:
        given_idx += 1
    else:
        given_idx += 2

    for idx, row in enumerate(sheet.iter_rows(min_row=given_idx, values_only=True)):
        hex_color = row[1] # row : list, [Colour]
        idx_row = given_idx+idx

        if hex_color:
            fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

            # Excel starts on indexed-1 and pandas is indexed-0 (Also header on the first line)
            # Fills only Column B
            # Removing Colour list
            cell_colour = sheet[idx_row][1] # Colour, the cell value
            cell_spending = sheet[idx_row][3] # Income/Spending, the cell value
            cell_total = sheet[idx_row][4] # Total, the cell value

            cell_colour.fill = fill
            cell_colour.value = None

            # Setting to currency format
            cell_spending.number_format =  '£#,##0.00'
            cell_total.number_format = '£#,##0.00'

            if cell_spending.value < 0:
                cell_spending.font = Font(color="ff0000", name='Calibri')
            cell_spending.alignment = Alignment(horizontal='center')
            cell_total.alignment = Alignment(horizontal='center')

    # Save the workbook
    workbook.save(output_file)
    workbook.close()

def clean_year(file: str) -> None:
    """
    Goes through file (year) and cleans any missing memo changes. Will also change the colour.
    Output is printing in terminal for any missing memos that haven't been added in JSON file.

    The function will:
    - Go through a finance file (i.e. 2024.xlsx)
    - Read memo.json
    """

    with open('memo.json', 'r', encoding='utf-8') as f:
        memo_file = json.load(f)
    workbook = load_workbook(filename=file)

    avoid_memos = list(memo_file.keys())
    for month in workbook:

        # Start at first cell (execel is a start base of 1 and the first row is the column names)
        for idx, row in enumerate(month.iter_rows(min_row=2, values_only=True)):
            if not row[2] in avoid_memos:
                category, hex_color = memo_extraction(memo_file, row[2])

                if hex_color != "ffffff":
                    cell_colour = month[idx+2][1] # Colour, cell value
                    cell_category = month[idx+2][2]

                    fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                    cell_colour.fill = fill
                    cell_category.value = category

    # Save the workbook
    workbook.save(file)
    workbook.close()

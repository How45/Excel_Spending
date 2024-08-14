from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from helper_function import memo_extraction, rgb_to_rbga

class CleanFileManager:
    """
    Cleans files, changes anything that wasn't inserted on first input
    """

    def __init__(self, year) -> None:
        self.year = year
        self.month = None
        self.output_file = f'finance/{year}.xlsx'

    def clean_what_column(self) -> None:
        """
        Goes through What? Column as changes unreadable text to category
        """
        pass

    def clean_colour_column(self) -> None:
        """
        Clean colour column to RGB text to PatternFill
        """
        pass

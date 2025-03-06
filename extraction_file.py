"""
File to get statements to put them on spreadsheet
"""
import os
import json
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import Cell
import extract_info as extract
from clean_sheet import clean_cells

# STARTING_VALUE: int = 11774.72
class FinancialManager:
    """
    Deals with cleaning of statements and puts them in sheets
    """
    def __init__(self, bank, year, month, dir_file, starting_value) -> None:
        folder_name: str = dir_file.split('\\')[-1]

        self.bank:str = bank
        self.year:str = year
        self.month:str = month
        self.output_file:str = f'{dir_file}\\finance\\{year}.xlsx'
        self.memo_file:str = f'{dir_file}\\{folder_name}_memo.json'
        self.banks_file:str = f'{dir_file}\\{folder_name}_bank.json'
        self.starting_value:int = starting_value

    def _check_existing_bank(self) -> bool:
        """
        Checks if there's already the name of the bank in the month of that year
        """
        df = pd.read_excel(self.output_file, sheet_name=self.month)
        return self.bank in df['Bank'].values

    # Needs to be re-done (normalise so it can gt all data)
    def clean(self, statement: str) -> pd.DataFrame:
        """
        Cleans bank statement
        """
        with open(self.banks_file, 'r', encoding='utf-8') as f:
            bank_operations = json.load(f)

        bank_info = bank_operations.get(self.bank)
        if not bank_info:
            raise KeyError(f'No banks under {self.bank} found')

        data = pd.read_csv(statement)

        # Change up if there.s null values ATM banks i've used don't
        amount = data[bank_info['amount_column']].tolist()
        date = data[bank_info['Dates']].tolist()

        memo = [name.replace('\t', '').replace(' ', '')
                for name in data[bank_info['memo_column']].tolist()]

        # Removes any amount and memo
        with open(self.memo_file, 'r', encoding='utf-8') as f:
            memo_json = json.load(f)

        amounts, memos, colours, dates = [], [], [], []
        for index, key in enumerate(memo):
            if key.lower() not in [x.lower() for x in bank_info['remove_memo']]:
                type_memo, colour = extract.category_memos(memo_json, key)

                if colour == "ffffff":
                    print(f'❌ You need to add {type_memo} to json file')

                dates.append(date[index])
                memos.append(type_memo)
                colours.append(colour)
                amounts.append(amount[index])

        last_total_cell = self._last_cell_amount()

        return self._to_dataframe(amounts, memos, last_total_cell, colours, dates)

    def tally_account(self, df: pd.DataFrame) -> None:
        """
        Adds either a new year file or a new month sheet to the year
        """
        if not os.path.exists(self.output_file):
            df.to_excel(self.output_file, sheet_name=self.month, index=False)
            clean_cells(self.output_file, self.month, 0)

        elif os.path.exists(self.output_file):
            workbook = load_workbook(filename=self.output_file)
            sheets = workbook.sheetnames

            if self.month in sheets:
                if not self.check_existing_bank():
                    sheet = workbook[self.month]
                    last_row = sheet.max_row

                    with pd.ExcelWriter(self.output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer: # pylint: disable=abstract-class-instantiated
                        df.to_excel(writer, sheet_name=self.month, startrow=last_row, index=False, header=False)

                    clean_cells(self.output_file, self.month, int(last_row))

                else:
                    print(f'❗️ {self.bank} exists already in that year')
                    print(f'❗️ {self.bank}_{self.month}_{self.year}, can be removed')

            # New month in workbook
            else:
                with pd.ExcelWriter(self.output_file, engine='openpyxl', mode='a') as writer: # pylint: disable=abstract-class-instantiated
                    df.to_excel(writer, sheet_name=self.month, index=False)

                clean_cells(self.output_file, self.month, 0)
        else:
            raise ValueError('No files found')

    def update_sheets(self) -> None:
        """
        Reformats sheets/workbooks infront of it like referencing the last_cell_amount
        """

        fiance_workbooks: list[str] = os.listdir(path='finance/')
        years_ahead: list[str] = [files.split('.')[0] for files in fiance_workbooks if int(files.split('.')[0]) >= int(self.year)]
        closest_month: str = None
        closest_year: str = None

        if self.month == "12":
            try:
                closest_year = years_ahead[1]
                workbook = load_workbook(f'finance/{closest_year}.xlsx')
                workbook_sheets: list[str] = workbook.sheetnames
                workbook.close()

                workbook_sheets.sort()
                closest_month = workbook_sheets[0]
            except IndexError:
                # Should be because no year has been created ahead of it
                # This is in the case where you are adding new files in order
                pass

        else:
            for pointer_year in years_ahead:
                workbook = load_workbook(f'finance/{pointer_year}.xlsx')
                workbook_sheets = workbook.sheetnames
                workbook.close()

                closest_month = extract.closest_month_from_self(self.month, workbook_sheets)
                if closest_month:
                    closest_year = pointer_year
                    break

        if closest_month and closest_year:
                self._update_first_line(closest_month, closest_year)
        else:
            print(f"❌ Nothing to update -> {self.month}/{self.year}")

    def _update_first_line(self, page_no: str, closest_year: str) -> None:
        """
        Updates the closest workbook top row, linking to new last row
        """
        current_total_cell = self._last_cell_amount()

        workbook = load_workbook(filename=f'finance/{closest_year}.xlsx')
        sheet = workbook[page_no]

        # MIGHT NOT NEED THIS V
        if not current_total_cell[1]:
            function_total = f"=SUM(D2,{STARTING_VALUE})"

        elif current_total_cell[1] == f'finance/{closest_year}.xlsx':
            function_total = f"=SUM(D2,'{current_total_cell[2]}'!{current_total_cell[0].coordinate})"

        else:
            path_year = os.path.abspath(current_total_cell[1]).replace("\\","/")
            function_total = f"=SUM(D2,'file:///[{path_year}]{current_total_cell[2]}'!{current_total_cell[0].coordinate})"

        sheet.cell(row=2,column=5).value = function_total
        workbook.save(f'finance/{closest_year}.xlsx')
        workbook.close()

    def _to_dataframe(self, amounts: list[int], memos: list[str],
                  last_total_cell: tuple[Cell|None, str|None, str|None],
                  colours: list[int], dates: list[str]) -> pd.DataFrame:
        """
        Creates dataframe to a new sheet
        """
        df = pd.DataFrame()
        # Default value
        if last_total_cell[1] is None:
            cell_value: int = last_total_cell[0]
            cell_coordinate: str = "E2"
        else:
            cell_value: int = 0
            cell_coordinate: str = last_total_cell[0].coordinate

        iteration: int = 0
        for amount, memo, colour, date in zip(amounts, memos, colours, dates):
            if iteration == 0:
                # If we are adding a new bank in the same month. Iteration: 1 is skipped
                function_total, iteration = self._sum_function(last_total_cell[1], last_total_cell[2], iteration, cell_value, cell_coordinate)

            elif iteration == 1:
                cell_coordinate = "E2"
                function_total = f"=SUM(D3,{cell_coordinate})"

            else:
                cell_coordinate = f"E{int(cell_coordinate[1:])+1}"
                function_total = f"=SUM(D{int(cell_coordinate[1:])+1},{cell_coordinate})"

            row = {"Date": date,
                    "Colour": colour,
                    "What?": memo,
                    "Income/Spending": amount,
                    "Total": function_total,
                    "Bank": self.bank}
            df = df._append(row, ignore_index=True)
            iteration += 1
        return df

    def _last_cell_amount(self) -> tuple[Cell|None, str|None, str|None]:
        """
        Retrieves the last cell closest to current month/year. First month then year.
        """
        # If the year before exits
        year_before_dir: str = f'finance/{extract.privious_year_from_self(self.year)}.xlsx'

        # This year
        if os.path.exists(self.output_file):
            workbook: Workbook = load_workbook(filename=self.output_file)
            sheets: list[str] = workbook.sheetnames

            # Current month
            if self.month in sheets:
                sheet = workbook[self.month]

                last_cell = extract.last_row(sheet)
                workbook.close()
                return (last_cell,self.output_file, self.month)

            privious_month: str = '0'+str(int(self.month)-1) if int(self.month)-1 < 10 else str(int(self.month)-1)

            # Privious month
            if privious_month in sheets:
                sheet = workbook[privious_month]
            else:
                print('❗️ no privious month found')
                return (STARTING_VALUE, None, None)

            last_cell = extract.last_row(sheet)
            workbook.close()
            return (last_cell, self.output_file, privious_month)

        # Last Year
        elif os.path.exists(year_before_dir):
            workbook = load_workbook(filename=year_before_dir)
            sheets = workbook.sheetnames

            try:
                sheet_str = sheets[-1]
                sheet = workbook[sheet_str]
                # Gets the last month in that year, returning last total of that month
                last_cell = extract.last_row(sheet)
                workbook.close()
                return (last_cell, year_before_dir, sheet_str)
            except ValueError as val_error:
                print(f"❌ {val_error}")

        # None
        else:
            print('❗️ no privious year has been found')
            return (STARTING_VALUE, None, None)

    def _sum_function(self, last_year: str|None, last_month: str|None, iteration: int,
                     cell_value: int, cell_coordinate: str) -> str:
        """
        Computes the last total for the first row of the sheet.

        The function determines how the first row should derive the last total, either by:
        - Continuing from the same month
        - Fetching the total from another sheet
        - Fetching the total from another workbook

        Parameters:
        ----------
        last_year : str|None
            The privious year gotten from the last_total_cell or None
        last_month : str
            The privious month gotten from the last_total_cell
        iteration : int
            The row number for the current Income/Spending being processed.
            Value will be +2 as excel starts on base 1 and we are skipping the header row
        cell_value : int
            A starting value or 0 if no value is provided.
        cell_coordinate : str
            The cell coordinate indicating where the total is located (e.g. 'E2', 'E36').
        """
        # First time creating a workbook ever
        if last_year is None:
            function_total = f'=SUM(D2,{cell_value})'

        # Same Year
        elif last_year == self.output_file:
            # Same Month
            if last_month == self.month:
                # +1 as row starts on 2
                function_total = f"=SUM(D{int(cell_coordinate[1:])+1},{cell_coordinate})"
                iteration += 1

            # Different Month in Workbook
            else:
                function_total = f"=SUM(D2,'{last_month}'!{cell_coordinate})"

        # LibreOffice Style
        else:
            path_year = os.path.abspath(last_year).replace("\\","/")

            function_total = f"=SUM(D2,'file:///[{path_year}]{last_month}'!{cell_coordinate})"

        return function_total, iteration

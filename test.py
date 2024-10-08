"""
.
"""
import unittest
import extract_info as hf
import extraction_file as ef
from openpyxl import load_workbook

class TestExtractionFile(unittest.TestCase):
    """
    .
    """
    def get_last_total(self):
        """
        .
        """
        file_name = 'statements/FirstD_12_2023.csv'
        bank, month, year = hf.extract_name(file_name.split('.')[0])

        statement = ef.FinanacialManager(bank, year, month)
        print(statement.get_last_total())

    def get_last_cell(self):
        """
        .
        """
        workbook = load_workbook(filename='finance/2023.xlsx')
        print(workbook.sheetnames)
        sheet = workbook['11']

        for row in sheet.iter_rows(min_row=sheet.max_row-1, max_row=sheet.max_row-1, min_col= 5, max_col= 5, values_only=False):
            for cell in row:
                workbook.close()
                # Returns the sheet the cell is from, the cell value,
                # the value in that cell,the whole cell tuple, veriable type
                print(cell.parent.title, cell.coordinate, cell.value, cell, type(cell)) # String

    def close_file(self):
        """
        .
        """
        for file_name in ['statements/Barclay_11_2023.csv',
                      'statements/Barclay_12_2023.csv']:
            # file_name = 'statements/TestA_11_2023.csv'
            bank, month, year = hf.extract_name(file_name.split('.')[0])

            statement = ef.FinanacialManager(bank, year, month)
            data = statement.clean(file_name)
            statement.tally_account(data)

        # Test the update
        file_name = 'statements/TestA_11_2023.csv'
        bank, month, year = hf.extract_name(file_name.split('.')[0])

        statement = ef.FinanacialManager(bank, year, month)
        data = statement.clean(file_name)
        statement.tally_account(data)
        statement.update_sheets()

if __name__ == '__main__':
    unittest.main()
